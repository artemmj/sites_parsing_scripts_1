from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium import webdriver
from selenium.common.exceptions import WebDriverException


# Настроить файл excel
wb = Workbook()
ws = wb.active
ws['A1'] = 'Название'
ws['B1'] = 'Цена'
ws['C1'] = 'URL изображения'
row_pointer = 2


url = 'https://bluehawaii.ru'
# Получить все категории
driver = webdriver.Chrome()
url_categories = []
try:
    driver.get(url)
    html = driver.page_source
    soup = BeautifulSoup(html, 'lxml')
    result = soup.find_all('a', attrs={'class': 't-menu__link-item'})
    for idx, r in enumerate(result):
        if idx == 0:
            url_categories.append(r['href'])
            continue
        url_categories.append(url + r['href'])
    url_categories = url_categories[:6]
except WebDriverException as e:
    print({'Selenium ERROR': str(e)})
    del driver
    driver = webdriver.Chrome()


for url in url_categories:
    ws.cell(row=row_pointer, column=1, value=url)
    row_pointer += 1

    print(f'({url}) Загрузка страницы...')
    driver.get(url)
    html = driver.page_source
    soup = BeautifulSoup(html, 'lxml')

    products = soup.find_all('div', attrs={'class': 'js-product'})

    for product in products:
        name = product.find('div', attrs={'class': 'js-store-prod-name'})
        name = name.contents[0] if name.contents else None

        price = product.find('div', attrs={'class': 'js-product-price'})
        price = price.contents[0] if price.contents else None

        img = product.find('div', attrs={'class': 'js-product-img'})
        img_url = img['data-original'] if img else None

        ws.cell(row=row_pointer, column=1, value=name)
        ws.cell(row=row_pointer, column=2, value=price)
        ws.cell(row=row_pointer, column=3, value=img_url)
        row_pointer += 1

driver.close()
wb.save('bluehawaii.xlsx')
print('Успешно завершено')
