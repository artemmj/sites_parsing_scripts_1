from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium import webdriver
from selenium.common.exceptions import WebDriverException


# Настроить файл excel
wb = Workbook()
ws = wb.active
ws['A1'] = 'Название'
ws['B1'] = 'Цена'
ws['C1'] = 'URL изображения'
row_pointer = 2


url = 'https://aptekasy.ru/stolitsy'
# Получить все категории
driver = webdriver.Chrome()
url_categories = []
try:
    driver.get(url)
    html = driver.page_source
    soup = BeautifulSoup(html, 'lxml')
    result = soup.find_all('a', attrs={'class': 'button'})
    for r in result:
        url_categories.append(url + r['href'])
    url_categories = url_categories[3:]
except WebDriverException as e:
    print({'Selenium ERROR': str(e)})
    del driver
    driver = webdriver.Chrome()


for url in url_categories:
    ws.cell(row=row_pointer, column=1, value=url)
    row_pointer += 1

    print(f'({url}) Загрузка страницы...')
    driver.get(url)
    html = driver.page_source
    soup = BeautifulSoup(html, 'lxml')

    products = soup.find_all('div', attrs={'class': 'product-item2'})

    for product in products:
        name = product.find('div', attrs={'class': 'product-list2'})
        name = name.h3.contents[0]
        price = product.find('span', attrs={'class': 'price2'})
        price = float(price.contents[0][2:])
        img_url = product.find('img')
        img_url = url + img_url['src']

        ws.cell(row=row_pointer, column=1, value=name)
        ws.cell(row=row_pointer, column=2, value=price)
        ws.cell(row=row_pointer, column=3, value=img_url)
        row_pointer += 1

driver.close()
wb.save('aptekasy.xlsx')
print('Успешно завершено')
