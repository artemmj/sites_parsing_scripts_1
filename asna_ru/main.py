import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook


def get_html(url):
    ''' Получить страницу '''
    r = requests.get(url)
    # r.encoding = 'utf8'
    return r.text


# Настроим файл excel
wb = Workbook()
ws = wb.active
ws['A1'] = 'Название'
ws['B1'] = 'Цена'
ws['C1'] = 'URL изображения'

row_pointer = 2


for page in range(1, 900):
    print(f'Загрузка страницы {page}...')
    html = get_html(f'https://www.asna.ru/catalog/lekarstva_i_bady/?PAGEN_1={page}')

    print('Парсинг...')
    soup = BeautifulSoup(html, 'lxml')
    result = soup.find_all('div', attrs={'class': 'product'})

    for el in result:

        # URL картинки
        prod_image = el.find('div', attrs={'class': 'product__image'})
        img_url = prod_image.find('img')['data-src']

        # Название
        prod_info = el.find('div', attrs={'class': 'product__information'})
        title_blk = prod_info.find('p', attrs={'class': 'product__title'})
        name = title_blk.contents[0]

        # Цена
        prod_info = el.find('div', attrs={'class': 'product__buy'})
        price_blk = prod_info.find('link', attrs={'itemprop': 'price'})
        price = price_blk['content'] if price_blk else None

        ws.cell(row=row_pointer, column=1, value=name)
        ws.cell(row=row_pointer, column=2, value=price)
        ws.cell(row=row_pointer, column=3, value=img_url)
        row_pointer += 1

    print(f'Записана страница: {page}\n')

wb.save('asnaru.xlsx')
print('Успешно завершено')
