import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook


def get_html(url):
    ''' Получить страницу '''
    r = requests.get(url)
    # r.encoding = 'utf8'
    return r.text


# Настроим файл excel
wb = Workbook()
ws = wb.active
ws['A1'] = 'Название'
ws['B1'] = 'Цена'
ws['C1'] = 'URL изображения'

row_pointer = 2

animals = {
    'dogs': 253, 'cats': 98, 'vetapteka': 40,
    'birds': 15, 'akvariumistika': 67, 'rodents': 30,
}

for animal, cnt in animals.items():

    ws.cell(row=row_pointer, column=1, value=animal)
    row_pointer += 1

    for page in range(1, cnt):
        print(f'Загрузка страницы {page} группы {animal}...')
        html = get_html(f'https://www.bethowen.ru/catalogue/{animal}/?PAGEN_1={page}')

        print('Парсинг...')
        soup = BeautifulSoup(html, 'lxml')
        result = soup.find_all('div', attrs={'itemprop': 'itemListElement'})

        for el in result:

            # URL картинки
            prod_image = el.find('img', attrs={'itemprop': 'image'})
            img_url = prod_image['src'] if prod_image else None

            # Название
            name = prod_image['title'] if prod_image else None

            # Цена
            prod_info = el.find('meta', attrs={'itemprop': 'price'})
            price = prod_info['content'] if prod_info else None

            ws.cell(row=row_pointer, column=1, value=name)
            ws.cell(row=row_pointer, column=2, value=price)
            ws.cell(row=row_pointer, column=3, value=img_url)
            row_pointer += 1

        print(f'Записана страница: {page} группы {animal}\n')

wb.save('bethowen.xlsx')
print('Успешно завершено')
