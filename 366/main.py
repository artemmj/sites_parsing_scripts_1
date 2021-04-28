import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook


def get_html(url):
    ''' Получить страницу '''
    r = requests.get(url)
    # r.encoding = 'utf8'
    return r.text


# Настроим файл excel
wb = Workbook()
ws = wb.active
ws['A1'] = 'Название'
ws['B1'] = 'Цена'
ws['C1'] = 'URL изображения'

row_pointer = 2

categories = {
    'lekarstva': 76,
    'vitaminy-i-bady': 17,
    'kosmetika': 32,
    'gigiena': 20,
    'medicinskie-izdelija-i-pribory': 11,
    'mama-i-malysh': 11,
    'dieticheskoe-pitanie': 3,
    'ukhod-za-bolnymi': 4,
    'ortopedija': 17,
}

for category, cnt in categories.items():

    ws.cell(row=row_pointer, column=1, value=category)
    row_pointer += 1

    for page in range(1, cnt):
        print(f'Загрузка страницы {page} группы {category}...')
        html = get_html(f'https://366.ru/c/{category}/?page={page}')

        print('Парсинг...')
        soup = BeautifulSoup(html, 'lxml')
        result = soup.find_all('div', attrs={'class': 'c-prod-item c-prod-item--grid'})

        for el in result:

            # URL картинки
            prod_image = el.find('img', attrs={'itemprop': 'url'})
            img_url = prod_image['src'] if prod_image else None

            # Название
            name = prod_image['alt'] if prod_image else None

            # Цена
            prod_info = el.find('meta', attrs={'itemprop': 'price'})
            price = prod_info['content'] if prod_info else None

            ws.cell(row=row_pointer, column=1, value=name)
            ws.cell(row=row_pointer, column=2, value=price)
            ws.cell(row=row_pointer, column=3, value=img_url)
            row_pointer += 1

        print(f'Записана страница: {page} группы {category}\n')

wb.save('366.xlsx')
print('Успешно завершено')
