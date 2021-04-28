from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium import webdriver
from selenium.common.exceptions import WebDriverException

# Настроить файл excel
wb = Workbook()
ws = wb.active
ws['A1'] = 'Название'
ws['B1'] = 'Цена'
ws['C1'] = 'URL изображения'
row_pointer = 2


url = 'https://zdravcity.ru'
# Получить все категории
driver = webdriver.Chrome()
url_categories = []
try:
    driver.get(url)
    html = driver.page_source
    soup = BeautifulSoup(html, 'lxml')
    categories = soup.find_all('li', attrs={'class': 'b-header-navigation-new__item'})  # noqa
    categories = categories[:8]

    for category in categories:
        subcats = category.find_all('a', attrs={'class': 'b-header-navigation-new__link'})  # noqa
        for subcat in subcats:
            url_categories.append(subcat['href'])

except WebDriverException as e:
    print({'Selenium ERROR': str(e)})
    del driver
    driver = webdriver.Chrome()


# Сколько категорий
length = len(url_categories)

for idx, url in enumerate(url_categories):
    ws.cell(row=row_pointer, column=1, value=url)
    row_pointer += 1

    for page in range(1, 9999):
        print(f'({url} ({idx+1} из {length}))\nЗагрузка и обработка страницы {page}...')  # noqa
        url_with_page = url + f'?PAGEN_1={page}'
        driver.get(url_with_page)
        html = driver.page_source
        soup = BeautifulSoup(html, 'lxml')

        result = soup.find_all('div', attrs={'class': 'b-product-item-new__wrapper'})  # noqa
        if len(result) > 1:
            for prod in result:
                name = prod.find('a', attrs={'class': 'b-product-item-new__title'})  # noqa
                if name:
                    name = name.contents[0]
                    name = name.strip()
                else:
                    name = None

                price = prod.find('span', attrs={'class': 'b-product-item-new__price--new-no-wrap'})  # noqa
                if price:
                    price = price.find('span')
                    if price:
                        price = price.contents[0]
                        price = price.strip()
                    else:
                        price = None
                else:
                    price = None

                img = prod.find('img', attrs={'type': 'image/jpeg'})  # noqa
                img_url = ('https://zdravcity.ru' + img['data-src']) if img else None  # noqa

                ws.cell(row=row_pointer, column=1, value=name)
                ws.cell(row=row_pointer, column=2, value=price)
                ws.cell(row=row_pointer, column=3, value=img_url)
                row_pointer += 1
        else:
            break


driver.close()
wb.save('zdravcity.xlsx')
print('Успешно завершено')
