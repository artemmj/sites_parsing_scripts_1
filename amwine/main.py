import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium import webdriver
from selenium.common.exceptions import WebDriverException


def get_html(url):
    ''' Получить страницу '''
    r = requests.get(url)
    r.encoding = 'utf8'
    return r.text


# Настроим файл excel
wb = Workbook()
ws = wb.active
ws['A1'] = 'Название'
ws['B1'] = 'Цена'
ws['C1'] = 'URL изображения'

row_pointer = 2

categories = {
    'vino': 458,
    'igristoe_vino_i_shampanskoe': 62,
    'krepkie_napitki': 171,
    'pivo': 27,
    'sidr': 5,
    'voda_i_soki': 28,
    'aksessuary': 27,
    'product': 86,
    'certificates_and_discount_cards': 2,
    'am-collection': 171
}

driver = webdriver.Chrome()


for category, cnt in categories.items():
    ws.cell(row=row_pointer, column=1, value=category)
    row_pointer += 1

    for page in range(1, cnt):
        try:
            print(f'({category}) Загрузка страницы {page}... ({cnt} всего)')
            # html = get_html(f'https://amwine.ru/catalog/{category}/?page={page}')  # noqa
            driver.get(f'https://amwine.ru/catalog/{category}/?page={page}')
            html = driver.page_source
            print('Парсинг...')
            soup = BeautifulSoup(html, 'lxml')
                                                # class    catalog-list-item articles-selector js-catalog-item     # noqa
            result = soup.find_all('div', attrs={'class': 'catalog-list-item articles-selector js-catalog-item'})  # noqa

            for el in result:

                try:
                    name = el['data-name']
                except KeyError:
                    name = None

                try:
                    price = el['data-price']
                except KeyError:
                    price = None

                try:
                    ell = el.find('img')
                    img_url = 'https://amwine.ru' + ell['data-src']
                except Exception:
                    img_url = None

                ws.cell(row=row_pointer, column=1, value=name)
                ws.cell(row=row_pointer, column=2, value=price)
                ws.cell(row=row_pointer, column=3, value=img_url)
                row_pointer += 1

        except WebDriverException as e:
            print({'Selenium ERROR': str(e)})
            del driver
            driver = webdriver.Chrome()

        print(f'Записана страница: {page} группы {category}\n')

wb.save('amwine.xlsx')
driver.close()
print('Успешно завершено')
