from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium import webdriver
from selenium.common.exceptions import WebDriverException


# Настроить файл excel
wb = Workbook()
ws = wb.active
ws['A1'] = 'Название'
ws['B1'] = 'Цена'
ws['C1'] = 'URL изображения'
row_pointer = 2


# Получить все категории
driver = webdriver.Chrome()
url_categories = []
try:
    driver.get('https://vkusvill.ru/goods/')
    html = driver.page_source
    soup = BeautifulSoup(html, 'lxml')
    result = soup.find_all('a', attrs={'class': 'VVCatalog2020Menu__Link'})
    for r in result:
        url_categories.append('https://vkusvill.ru' + r['href'])
    url_categories = url_categories[3:]
except WebDriverException as e:
    print({'Selenium ERROR': str(e)})
    del driver
    driver = webdriver.Chrome()


for url in url_categories:
    ws.cell(row=row_pointer, column=1, value=url)
    row_pointer += 1

    for page in range(1, 9999):
        print(f'({url}) Загрузка страницы {page}...')
        url_with_page = url + f'?PAGEN_1={page}'
        driver.get(url_with_page)
        html = driver.page_source
        soup = BeautifulSoup(html, 'lxml')

        result = soup.find_all('div', attrs={'class': 'ProductCards__item'})  # noqa
        if len(result) > 1:
            for prod in result:
                name = prod.find('a', attrs={'class': 'ProductCard__link js-datalayer-catalog-list-name'})  # noqa
                name = name['title'] if name else None
                price = prod.find('span', attrs={'class': 'Price__value'})
                price = price.contents[0] if price else None
                img = prod.find('img', attrs={'class': 'ProductCard__imageImg lazyload'})  # noqa
                img_url = ('https://vkusvill.ru' + img['data-src']) if img else None  # noqa

                ws.cell(row=row_pointer, column=1, value=name)
                ws.cell(row=row_pointer, column=2, value=price)
                ws.cell(row=row_pointer, column=3, value=img_url)
                row_pointer += 1
        else:
            break

driver.close()
wb.save('vkusvill.xlsx')
print('Успешно завершено')
